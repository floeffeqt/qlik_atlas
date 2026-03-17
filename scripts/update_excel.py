"""Add Security Fixes and Git Integration Plan worksheets to the analytics catalog."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("qlik_atlas_analytics_catalog.xlsx")

# ── Styles ──
header_font_white = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def status_fill(status):
    s = (status or "").lower()
    if s in ("done", "erledigt", "completed"):
        return green_fill
    elif s in ("open", "offen", "pending"):
        return red_fill
    elif s in ("in_progress", "in arbeit"):
        return yellow_fill
    return None


def write_sheet(ws, headers, data, widths):
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    status_col = None
    for idx, h in enumerate(headers):
        if h.lower() == "status":
            status_col = idx
            break

    for r, row_data in enumerate(data, 2):
        fill = status_fill(row_data[status_col]) if status_col is not None else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap
            cell.border = thin_border
            if fill:
                cell.fill = fill

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════
# Sheet 1: Security Fixes
# ══════════════════════════════════════════
if "Security Fixes" in wb.sheetnames:
    del wb["Security Fixes"]

ws_sec = wb.create_sheet("Security Fixes")

sec_headers = [
    "ID", "Prioritaet", "Kategorie", "Titel", "Beschreibung",
    "Status", "Umsetzungsdatum", "Betroffene Dateien", "Hinweise",
]

sec_data = [
    ["K1", "Kritisch", "Secrets",
     ".env Secrets aus Git-History entfernen",
     "Produktive Secrets (.env) befinden sich in der Git-History. BFG Repo-Cleaner oder git filter-repo notwendig.",
     "Offen", "",
     ".env, .gitignore",
     "Erfordert Force-Push + Credential-Rotation"],
    ["K2", "Kritisch", "Autorisierung/RLS",
     "RLS Policies: Customer-Access-Check",
     "Alle 17 project-scoped Tabellen hatten Policies ohne app_has_customer_access() Pruefung.",
     "Erledigt", "2026-03-13",
     "backend/alembic/versions/0019_rls_customer_access.py",
     "Migration 0019"],
    ["K3", "Kritisch", "DB Pooling",
     "Explizite Connection Pool Settings",
     "SQLAlchemy Engine hatte keine expliziten pool_size/max_overflow/pool_recycle Settings.",
     "Erledigt", "2026-03-13",
     "backend/app/database.py",
     "Env-Vars: DB_POOL_SIZE, DB_POOL_MAX_OVERFLOW, DB_POOL_RECYCLE_SECONDS"],
    ["H4", "Hoch", "CORS",
     "CORS Method/Header Whitelisting",
     "allow_methods und allow_headers waren auf Wildcard (*) gesetzt.",
     "Erledigt", "2026-03-14",
     "backend/main.py",
     "Explizite Methoden + Content-Type Header"],
    ["H5", "Hoch", "Security Headers",
     "HSTS + CSP Haertung",
     "Strict-Transport-Security fehlte. CSP unsafe-inline bei style-src bleibt dokumentierter Trade-off.",
     "Erledigt", "2026-03-14",
     "backend/shared/security_headers.py, frontend/nginx.conf",
     "HSTS: max-age=63072000"],
    ["H6", "Hoch", "Token Security",
     "HMAC-SHA256 Refresh Token Hashing",
     "Refresh Tokens waren mit einfachem SHA256 gehasht statt HMAC.",
     "Erledigt", "2026-03-14",
     "backend/app/auth/utils.py, backend/app/auth/routes.py",
     "7-Tage Uebergangsphase mit Dual-Hash Lookup"],
    ["H7", "Hoch", "Logging",
     "Structured Error Logging",
     "13+ Exception Handler hatten kein Logging. print() statt logger. Stille Except-Bloecke.",
     "Erledigt", "2026-03-14",
     "backend/main.py, backend/app/db_runtime_views.py, backend/fetchers/artifact_graph.py",
     "Logger: atlas.api, atlas.runtime, atlas.graph"],
    ["H8", "Hoch", "CI/CD",
     "Tests in CI/CD Pipeline",
     "Docker-basierte Test-Pipeline mit pytest. Multi-Stage Dockerfile (test Target). 12 bestehende Testdateien.",
     "Erledigt", "2026-03-17",
     "backend/Dockerfile, backend/requirements-test.txt, docker-compose.yml",
     "docker compose --profile test run --rm test"],
    ["M1", "Mittel", "Testing",
     "Backend Test Coverage erhoehen",
     "56% der Backend-Dateien haben keine Tests. 0% Frontend-Tests.",
     "Offen", "",
     "tests/",
     "Fokus auf Auth, Customers, Fetch Jobs"],
    ["M2", "Mittel", "Performance",
     "Fehlende DB Indexes",
     "Composite Indexes fuer haeufige Query-Patterns fehlen.",
     "Offen", "",
     "backend/alembic/versions/",
     "z.B. (project_id, app_id) auf lineage_nodes"],
    ["M3", "Mittel", "Performance",
     "Graph Pagination",
     "Voller In-Memory Load des Lineage-Graphen. Problematisch bei grossen Tenants.",
     "Offen", "",
     "backend/app/db_runtime_views.py",
     "Cursor-basierte Pagination empfohlen"],
]

write_sheet(ws_sec, sec_headers, sec_data, [6, 10, 18, 35, 60, 10, 14, 55, 45])


# ══════════════════════════════════════════
# Sheet 2: Git Integration Plan
# ══════════════════════════════════════════
if "Git Integration Plan" in wb.sheetnames:
    del wb["Git Integration Plan"]

ws_git = wb.create_sheet("Git Integration Plan")

git_headers = [
    "Phase", "Titel", "Beschreibung", "Lieferobjekte",
    "Status", "Abgeschlossen am", "Betroffene Dateien/Module",
    "Abhaengigkeiten", "Hinweise",
]

git_data = [
    ["P0", "Grundstruktur (Provider + DB)",
     "DB-Tabellen, Git Provider Interface, QlikClient POST, Customer Git-Felder",
     "Migration 0020\ngit_bridge/ Modul (5 Dateien)\nQlikClient post_json/get_text\nCustomer Model + Routes Update\n8 neue API-Endpoints",
     "Erledigt", "2026-03-16",
     "backend/alembic/versions/0020_script_sync_tables.py\nbackend/app/git_bridge/__init__.py\nbackend/app/git_bridge/provider.py\nbackend/app/git_bridge/github_provider.py\nbackend/app/git_bridge/gitlab_provider.py\nbackend/app/git_bridge/service.py\nbackend/app/git_bridge/routes.py\nbackend/shared/qlik_client.py\nbackend/app/models.py\nbackend/app/customers/routes.py\nbackend/main.py",
     "Keine",
     "Provider-Abstraction: GitHub + GitLab\nToken AES-256-GCM verschluesselt\nRLS Policies auf beiden neuen Tabellen"],
    ["P1", "Drift Detection + UI",
     "Drift-Status pro App, Frontend script-sync.html Admin-Seite",
     "Enriched /status + /overview Endpoints\nscript-sync.html (Mapping CRUD, Ampel, History)\nNav-Links auf allen 6 Seiten\natlas-shared.js Update",
     "Erledigt", "2026-03-16",
     "backend/app/git_bridge/routes.py\nfrontend/script-sync.html\nfrontend/assets/atlas-shared.js\nfrontend/*.html (Nav-Links)",
     "P0",
     "Script-Normalisierung: BOM, CRLF, Trailing WS\nBatch-Load App-Namen (kein N+1)"],
    ["P2", "Publish (Git -> Qlik)",
     "Script aus Git-Repo in Qlik App deployen mit Pre-Checks",
     "POST /{app_id}/publish\nGET /{app_id}/diff\nDeployment Audit-Log",
     "Offen", "",
     "backend/app/git_bridge/routes.py\nbackend/shared/qlik_client.py",
     "P1 + Qlik API Key Scopes",
     "POST /api/v1/apps/{appId}/scripts\nPre-Checks: RLS, Divergenz, Branch-Policy\nOptional: Reload-Trigger"],
    ["P3", "Reverse Sync (Qlik -> Git)",
     "Manuelle Qlik-Aenderungen zurueck ins Git committen",
     "POST /{app_id}/commit-to-git",
     "Offen", "",
     "backend/app/git_bridge/routes.py",
     "P2",
     "GitProvider.write_file() bereits implementiert"],
    ["P4", "Webhook Auto-Drift",
     "Bei Git-Push automatisch betroffene Apps als drift markieren",
     "POST /api/webhooks/git/{provider}\nWebhook-Secret Validierung",
     "Offen", "",
     "backend/app/git_bridge/webhooks.py (neu)",
     "P1 + oeffentlicher Endpoint",
     "GitHub: X-Hub-Signature-256\nGitLab: X-Gitlab-Token"],
    ["P5", "Erweiterungen",
     "Auto-Publish, Batch-Publish, Branch-Preview, Atlas CLI",
     "Optional: CLI Client, Auto-Deploy auf release/* Branches",
     "Offen", "",
     "TBD",
     "P2",
     "Optionale Features nach Bedarf"],
]

write_sheet(ws_git, git_headers, git_data, [7, 28, 55, 55, 12, 16, 55, 35, 50])

wb.save("qlik_atlas_analytics_catalog.xlsx")
print("OK: 2 sheets added -> Security Fixes, Git Integration Plan")